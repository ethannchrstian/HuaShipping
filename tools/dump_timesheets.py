"""Dump every voyage sheet from the Excel time sheets into one JSON reference file.

This is a *raw, faithful* dump used as ground truth while writing the planning
docs (ERD, calculation spec, import mapping). It deliberately does NOT clean or
fix the data — real data-entry errors must stay visible. The production import
script (later, TDD) will be derived from docs/05-import-mapping.md.

Usage:  python tools/dump_timesheets.py
Output: docs/reference/timesheets-raw.json
"""

from __future__ import annotations

import datetime as dt
import glob
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "docs" / "reference" / "timesheets-raw.json"

HEADER_LABELS = {
    "No. Kontrak": "contract_no",
    "Kwitansi Nomor": "invoice_no",
    "Muatan": "cargo",
    "Lokasi Muat": "load_location",
    "Lokasi Bongkar": "discharge_location",
    "Lama Muat/Bongkar": "laytime",
    "Demurrage": "demurrage",
}


def cell_repr(v):
    """Serialize a cell value faithfully, keeping type info visible."""
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        # Excel duration formulas >24h come back as fake 1900-era datetimes;
        # keep them distinguishable from real dates.
        if v.year < 1990:
            return {"excel_duration_artifact": v.isoformat()}
        return v.isoformat()
    if isinstance(v, dt.time):
        return v.isoformat(timespec="minutes")
    if isinstance(v, dt.timedelta):
        total_min = int(v.total_seconds() // 60)
        return {"timedelta": f"{total_min // 1440}d {(total_min % 1440) // 60:02d}:{total_min % 60:02d}"}
    return v


def parse_sheet(ws):
    sheet = {"sheet_name": ws.title, "title": ws["B2"].value, "header": {}, "rows": []}
    m = re.search(r"VOY\.?\s*(V\d+)", str(ws["B2"].value or ""), re.I)
    sheet["voyage_code"] = m.group(1) if m else None

    # Header block: label in column B, value in column C (rows 3..10)
    for r in range(3, 11):
        label = ws.cell(row=r, column=2).value
        if not label:
            continue
        label = str(label).strip()
        key = next((k for lbl, k in HEADER_LABELS.items() if label.startswith(lbl)), None)
        val = ws.cell(row=r, column=3).value
        if key:
            val = re.sub(r"^\s*:\s*", "", str(val)) if val is not None else None
            sheet["header"][key] = val
        else:
            sheet["header"].setdefault("_unrecognized", []).append({label: cell_repr(val)})

    # Table rows: from row 13 down to the signature block
    for r in range(13, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value  # total days
        d = ws.cell(row=r, column=4).value  # total hh:mm / timedelta
        if label is None and c is None and d is None:
            continue
        if label and re.search(r"Dibuat|Diketahui", str(label)):
            break
        text = str(label).strip() if label else None
        if text and text.lower().startswith("total"):
            kind = "subtotal"
        elif text and "prorata" in text.lower():
            kind = "laytime_contract"
        elif text and "demurrage" in text.lower():
            kind = "demurrage"
        elif text is None:
            kind = "grand_total"
        else:
            kind = "activity"
        row = {
            "excel_row": r,
            "kind": kind,
            "label": text,
            "total_days": cell_repr(c),
            "total_time": cell_repr(d),
            "from_location": cell_repr(ws.cell(row=r, column=5).value),
            "start_date": cell_repr(ws.cell(row=r, column=6).value),
            "start_time": cell_repr(ws.cell(row=r, column=7).value),
            "to_location": cell_repr(ws.cell(row=r, column=8).value),
            "end_date": cell_repr(ws.cell(row=r, column=9).value),
            "end_time": cell_repr(ws.cell(row=r, column=10).value),
        }
        if ws.max_column >= 11:
            extra = ws.cell(row=r, column=11).value
            if extra is not None:
                row["col_K"] = cell_repr(extra)
        sheet["rows"].append(row)
    return sheet


def main():
    result = []
    for path in sorted(glob.glob(str(ROOT / "*.xlsx"))):
        wb = openpyxl.load_workbook(path, data_only=True)
        for name in wb.sheetnames:
            parsed = parse_sheet(wb[name])
            parsed["workbook"] = Path(path).name
            result.append(parsed)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(result, indent=1, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(result)} voyage sheets -> {OUT.relative_to(ROOT)}")
    for s in result:
        acts = sum(1 for r in s["rows"] if r["kind"] == "activity")
        print(f"  {s['workbook']:55s} {s['sheet_name']:28s} activities={acts}")


if __name__ == "__main__":
    main()
